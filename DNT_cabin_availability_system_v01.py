import os.path
import requests
import pandas as pd
import numpy
import datetime as dt
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side

# This function creates a list of months to check from todays month
def get_next_months_as_list(numberOfMonths=10) -> list:
    d = dt.date.today()
    list_of_months = [(d.strftime("%Y"),d.strftime("%m"))]
    for i in range(1,numberOfMonths):
        next_month = d + relativedelta(months=i)
        list_of_months.append((next_month.strftime("%Y"), next_month.strftime("%m")))
    return list_of_months

# yeah. i did not really do this properly...
def getcell(col_nr, row_nr) -> str:
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    letter = ""
    i  = 0
    while col_nr >= len(letters):
        letter = letters[i]
        col_nr = col_nr - len(letters)
        i += 1

    letter += letters[col_nr]
    return letter + str(row_nr)

# Gah i hate that int(None) wont work...
def fix_max(value):

    if value is None:
        return 1

    try:
        return int(value)
    except:
        return 1
        print("wtf")

# Not really why i have the one above. ah. yeah. because it will have 1 if it fails.
def intF(str):
    if str == None:
        return 0
    else:
        try:
            return int(str)
        except:
            return 0


## General config
number_of_months_to_check_ahead = 3 # 1 = this month
output_base_filename = "Tilgjengelige_DNT_hytter"
output_location = r"C:\00_TEMP"
excel_sheet = "Tilgjengelig"

## controller config
controller_list = r"C:\00_GIT\DNT_cabin_availability_system\DNT_controller.xlsx"
sheet = "controller"

last_gathered = "Sist undersøkt"
Max_number_of_accommodations = "Antall mulige reservasjoner"
order_system_ID = 'bestillings_system_ID'
List_of_accommodation_IDs = "List_of_accomodation_ID"
cabin_name = "Navn"
link = "Lenke til bestillingsystem"


## API config:
#PS: not sure if there are several api's as there is an 8 here.
base_availability_url = r"https://ws.visbook.com/8/api/"


## Excel style config:
col_nr_name = 0
col_nr_link = 1
col_nr_last_gathered = 2
col_nr_max = 3
col_nr_start_acc = 4

row_nr_table_heading = 8
row_nr_title = 0

width_accommodation_col = 4
width_name_cell = 50
width_other_cell = 20
width_max = 13
width_last_gathered = 20
width_link = 17

height_accommodation_col = 70

color_all_available = "7eed8d"
color_some_available = "f7d692"
color_none_available = "f2b8b6"

heading_font_size = 15
height_heading_row = 40



df = pd.read_excel(controller_list,sheet_name=sheet)

months_to_check = get_next_months_as_list(number_of_months_to_check_ahead)

# gather the data
for index, row in df.iterrows():

    if row[cabin_name] == None or type(row[cabin_name]) == float:
        pass

    elif "H#" in row[cabin_name] or row[cabin_name] == None:
        pass

    elif "STOP" in row[cabin_name]:
        break
    elif row[List_of_accommodation_IDs]  == None or type(row[List_of_accommodation_IDs]) == float:
        pass
    elif row[List_of_accommodation_IDs] == "":
        pass

    else:
        print(row[cabin_name])
        list_of_accommodations = str(row[List_of_accommodation_IDs]).split(",")
        df.at[index, link] = f"https://reservations.visbook.com/{intF(row[order_system_ID])}"
        df.at[index, last_gathered] = dt.datetime.now()
        df.at[index, Max_number_of_accommodations] = len(list_of_accommodations)

        for month in months_to_check:
            for accommodation in list_of_accommodations:
                # Creating quarry
                try:
                    request_url = base_availability_url + f"{int(row[order_system_ID])}/availability/{int(accommodation)}/{month[0]}-{month[1]}"
                    print("\t", request_url)


                    # Getting data
                    resp = requests.get(url=request_url)
                    data = resp.json()
                    # Adding data to dataframe
                    for step in data['items']:
                        date = step['date'].replace("T00:00:00", "")  # just to shorten date in final product. no need for time.
                        if date in df:
                            if step['webProducts'][0]['availability']['available']:
                                if numpy.isnan(df.at[index, date]):
                                    df.at[index, date] = 1
                                else:
                                    df.at[index, date] = df.at[index, date] + 1
                            else:
                                pass

                        else:
                            if step['webProducts'][0]['availability']['available']:
                                df.at[index, date] = 1
                            else:
                                df.at[index, date] = 0
                except:
                    pass


df.drop(order_system_ID,1,inplace=True)
df.drop(List_of_accommodation_IDs,1,inplace=True)

# # export to excel
date = dt.datetime.now()
output_file = os.path.join(output_location,date.strftime("%Y-%m-%d") + "_" + output_base_filename + ".xlsx")
df.to_excel(output_file, sheet_name=excel_sheet, startrow = row_nr_table_heading - 1, index = False)










# Prettify excel document:
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# heading
ws['A1'] = 'Oversikt over tilgjengelige DNT-hytter'
ws['A1'].font = Font(size = "30", bold=True)
ws['A2'] = 'Sist oppdatert:'
ws['A2'].font = Font(bold=True)
ws['B2'] = date.strftime("%Y-%m-%d")
ws['A3'] = 'Kontaktinformasjon:'
ws['A3'].font = Font(bold=True)
ws['B3'] = 'andreas.p.lorentzen@gmail.com'
ws['A4'] = 'Generell informasjon:'
ws['A4'].font = Font(bold=True)
ws['B4'] = 'Dette dokumentet er genrert automatisk av Andreas P. Lorentzen for privat bruk. Det er ikke nødvendigvis helt oppdatert.'


# fix first cols
al = Alignment(horizontal='left', vertical='bottom', wrapText=True)
# Name
ws[getcell(col_nr_name, row_nr_table_heading)].alignment = al
ws.column_dimensions[getcell(col_nr_name, row_nr_table_heading)[:-1]].width = width_name_cell

#l ink
ws[getcell(col_nr_link, row_nr_table_heading)].alignment = al
ws.column_dimensions[getcell(col_nr_link, row_nr_table_heading)[:-1]].width = width_link

# Last gathered
ws[getcell(col_nr_last_gathered, row_nr_table_heading)].alignment = al
ws.column_dimensions[getcell(col_nr_last_gathered, row_nr_table_heading)[:-1]].width = width_last_gathered

# Max
ws[getcell(col_nr_max, row_nr_table_heading)].alignment = al
ws.column_dimensions[getcell(col_nr_max, row_nr_table_heading)[:-1]].width = width_max

# height of header row
ws.row_dimensions[row_nr_table_heading].height = height_accommodation_col

# orientation of dates
al = Alignment(textRotation=90, horizontal='center')
for i in range(col_nr_start_acc,ws.max_column):
    ws[getcell(i,row_nr_table_heading)].alignment = al
    ws.column_dimensions[getcell(i,row_nr_table_heading)[:-1]].width = width_accommodation_col

# fix "tilgjengelige reservasjoner"
ws[getcell(col_nr=col_nr_start_acc,row_nr=row_nr_table_heading-1)] = "Oversikt over antall ledige reservjasoner pr. dato"
ws[getcell(col_nr=col_nr_start_acc,row_nr=row_nr_table_heading-1)].font = Font(bold=True)
ws.merge_cells(getcell(col_nr=col_nr_start_acc,row_nr=row_nr_table_heading-1)+":"+getcell(col_nr=ws.max_column-1,row_nr=row_nr_table_heading-1))


# Fix hyperlinks
for i in range(row_nr_table_heading + 1,ws.max_row+1):
    cell = getcell(col_nr=col_nr_link,row_nr=i)
    if ws[getcell(col_nr=col_nr_last_gathered, row_nr=i)].value == None or type(
            ws[getcell(col_nr=col_nr_last_gathered, row_nr=i)].value) == float:
        pass
    elif "H#" in ws[getcell(col_nr=col_nr_name,row_nr=i)].value:
        pass
    else:
        if ws[cell].value != "":
            ws[cell].hyperlink = ws[cell].value
            ws[cell].value = "Lenke"
            ws[cell].style = "Hyperlink"





# Fix max
al = Alignment(horizontal='left', vertical='bottom', wrapText=True)
for i in range(row_nr_table_heading + 1,ws.max_row+1):
    cell = getcell(col_nr=col_nr_max,row_nr=i)
    ws[cell].alignment = al

# Fix colors and accomodation
none = PatternFill(start_color=color_none_available, end_color=color_none_available, fill_type = "solid")
some = PatternFill(start_color=color_some_available, end_color=color_some_available, fill_type = "solid")
all_available =  PatternFill(start_color=color_all_available, end_color=color_all_available, fill_type = "solid")
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
al = Alignment(horizontal='center', vertical='center')

for row in range(row_nr_table_heading+1,ws.max_row+1):
    max_available = fix_max(ws[getcell(col_nr=col_nr_max,row_nr=row)].value)


    if ws[getcell(col_nr=col_nr_name,row_nr=row)].value == None or type(ws[getcell(col_nr=col_nr_name,row_nr=row)].value) == float:
        pass
    elif "H#" in ws[getcell(col_nr=col_nr_name,row_nr=row)].value:
        pass
    elif ws[getcell(col_nr=col_nr_last_gathered,row_nr=row)].value == "" or ws[getcell(col_nr=col_nr_last_gathered,row_nr=row)].value == None:
        pass
    else:
        for i in range(col_nr_start_acc, ws.max_column):
            ws[getcell(col_nr=i,row_nr=row)].alignment = al
            cell_value = ws[getcell(col_nr=i,row_nr=row)].value
            #print("\t", cell_value)
            if cell_value is None or cell_value == 0:
                ws[getcell(i, row)].fill = none
                ws[getcell(i, row)] = 0

            elif int(cell_value) < int(max_available):
                ws[getcell(i, row)].fill = some

            elif int(cell_value) == int(max_available):
                ws[getcell(i, row)].fill = all_available

            ws[getcell(i, row)].border = thin_border


# Fix headings
al = Alignment(horizontal='left', vertical='bottom', wrapText=True)
font = Font(size=heading_font_size,bold=True)

for i in range(row_nr_table_heading + 1,ws.max_row+1):
    cell = getcell(col_nr=col_nr_name,row_nr=i)
    if ws[cell].value is None or type(ws[cell].value) == float:
        pass
    elif "H#" in ws[cell].value:
        ws[cell] = ws[cell].value[2:]
        ws[cell].alignment = al
        ws[cell].font = font
        ws.row_dimensions[i].height = height_heading_row


# save file
wb.save(filename=output_file)


#print(df)
# df.to_csv(path_or_buf=r"C:\00_TEMP\crawl_DNT_hytter_test2.csv",sep="|",encoding="utf-8")

